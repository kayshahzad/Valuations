"""
aletheia/tools/dcf_excel_exporter.py

Institutional DCF Excel Model — Aletheia Universe Export
=========================================================
Generates a full institutional-grade DCF workbook for any ticker
in the Aletheia universe using data from the pipeline report JSON.

5 sheets:
  DCF   — P&L projection, FCF build, TV methods, sensitivity tables
  NWC   — Working capital schedule (DSO/DIH/DPO driven)
  WACC  — Cost of capital build + sensitivity
  AS1   — Income statement & cash flow assumptions
  AS2   — Balance sheet assumptions

Usage:
    python3 -m aletheia.tools.dcf_excel_exporter --ticker MSFT
    python3 -m aletheia.tools.dcf_excel_exporter --ticker LLY MSFT BRK-B
    python3 -m aletheia.tools.dcf_excel_exporter --all
"""

from __future__ import annotations
import json
import math
import argparse
import sys
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORT_DIR = Path("valuation_data/serving/latest")
OUTPUT_DIR = Path("valuation_data/serving/latest")

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E5FA3"
LIGHT_BLUE  = "D9E2F3"
VERY_LIGHT  = "EEF3FB"
WHITE       = "FFFFFF"
GOLD        = "C9A84C"
SENS_GOLD   = "FFF2CC"
INPUT_BLUE  = "0000FF"
FORMULA_BLK = "000000"
LINK_GREEN  = "008000"
RED_DARK    = "8B0000"
GREY_LIGHT  = "F2F2F2"

# Number formats
FMT_DOLLAR  = '$#,##0;($#,##0);"-"'
FMT_DOLLAR2 = '$#,##0.00;($#,##0.00);"-"'
FMT_PCT1    = '0.0%;(0.0%);"-"'
FMT_MULT    = '0.0x;(0.0x);"-"'
FMT_DEC4    = '0.0000'
FMT_INT     = '#,##0'

def _fill(hex_c): return PatternFill("solid", start_color=hex_c, end_color=hex_c)
def _font(bold=False, color=FORMULA_BLK, size=9, italic=False, name="Arial"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)
def _input(): return _font(color=INPUT_BLUE)
def _link():  return _font(color=LINK_GREEN)
def _hdr(size=9): return _font(bold=True, color=WHITE, size=size)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center")
def _right():  return Alignment(horizontal="right",  vertical="center")
def _thin(): return Side(style="thin")
def _border(b=True, t=False): return Border(bottom=_thin() if b else Side(), top=_thin() if t else Side())

def _set(ws, row, col, value, fill=None, font=None, align=None, fmt=None, border=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    if fill:   c.fill   = fill
    if font:   c.font   = font
    if align:  c.alignment = align
    if fmt:    c.number_format = fmt
    if border: c.border = border
    return c

def _hdr_row(ws, row, text, col_start, col_end, fill=MID_BLUE, size=9):
    _set(ws, row, col_start, text, fill=_fill(fill), font=_hdr(size=size), align=_left())
    for c in range(col_start+1, col_end+1):
        _set(ws, row, c, None, fill=_fill(fill))
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)


# ─────────────────────────────────────────────────────────────────────────────
# Data extractor
# ─────────────────────────────────────────────────────────────────────────────

class ReportData:
    """Extracts and organises all DCF inputs from a pipeline report JSON."""

    def __init__(self, report: dict, ticker: str):
        self.ticker = ticker.upper()
        self.report = report

        # Sections
        self.p2  = report.get("4_valuation_synthesis", {}).get("phase2_valuation", {}) or {}
        self.ft  = report.get("2_financial_translation", {}) or {}
        self.cf  = self.ft.get("clean_financials", {}) or {}
        self.rat = self.ft.get("ratios", {}) or {}
        self.er  = report.get("1_economic_reality", {}) or {}
        self.it  = report.get("4_valuation_synthesis", {}).get("investment_thesis", {}) or {}
        self.ps  = self.it.get("pillar_scores", {}) or {}

        # WACC
        self.wacc      = self._s(self.p2.get("wacc"))     or 0.10
        self.beta      = self._s(self.p2.get("beta"))     or 1.0
        self.rf        = self._s(self.p2.get("risk_free_rate")) or 0.04
        self.erp       = 0.0662   # Damodaran ERP

        # Financials (latest year)
        self.revenue_bn   = self._s(self.cf.get("revenue_bn"))   or 0
        self.ebitda_bn    = self._s(self.cf.get("ebitda_bn"))    or 0
        self.fcf_bn       = self._s(self.cf.get("fcf_bn"))       or 0
        self.net_debt_bn  = self._s(self.cf.get("net_debt_bn"))  or 0
        self.gross_margin = self._s(self.rat.get("gross_margin_pct")) or 0
        self.fcf_margin   = self._s(self.rat.get("fcf_margin_pct"))  or 0
        self.roic         = self._s(self.rat.get("roic"))            or 0
        self.ebit_margin  = self._s(self.rat.get("ebit_margin_pct")) or 0

        # DCF scenarios
        dcf3 = self.p2.get("three_scenario_dcf", {}) or {}
        self.bear_iv  = self._s(dcf3.get("bear", {}).get("intrinsic_per_share")) or 0
        self.base_iv  = self._s(dcf3.get("base", {}).get("intrinsic_per_share")) or 0
        self.bull_iv  = self._s(dcf3.get("bull", {}).get("intrinsic_per_share")) or 0
        self.bear_mos = self._s(dcf3.get("bear", {}).get("margin_of_safety"))    or 0
        self.base_mos = self._s(dcf3.get("base", {}).get("margin_of_safety"))    or 0
        self.bull_mos = self._s(dcf3.get("bull", {}).get("margin_of_safety"))    or 0
        self.base_ev  = self._s(dcf3.get("base", {}).get("ev"))                  or 0

        # Market price (derived from base IV and MoS)
        if self.base_iv > 0 and self.base_mos != -1:
            self.market_price = self.base_iv / (1 + self.base_mos) if (1+self.base_mos) != 0 else 0
        else:
            self.market_price = 0

        # Reverse DCF
        rdcf = self.p2.get("reverse_dcf", {}) or {}
        self.hist_cagr = self._s(rdcf.get("historical_cagr")) or 0.05
        self.impl_cagr = self._s(rdcf.get("implied_cagr_10y")) or 0.05
        self.rdcf_signal = rdcf.get("signal", "")

        # Multiple decomposition
        md = self.p2.get("multiple_decomposition", {}) or {}
        self.market_ev_ebitda    = self._s(md.get("market_ev_ebitda"))    or 0
        self.justified_ev_ebitda = self._s(md.get("justified_ev_ebitda")) or 0
        self.premium_pct         = self._s(md.get("premium_pct"))         or 0
        self.roic_wacc_spread    = self._s(md.get("roic_wacc_spread"))    or 0
        self.value_creation      = md.get("value_creation", "")

        # Conviction
        self.conviction_score = self.it.get("conviction_score")
        self.position_tier    = self.ps.get("position_tier", "")
        self.lifecycle_stage  = self.ps.get("lifecycle_stage", "")

        # Shares (derived from equity value and share price)
        if self.market_price > 0 and self.base_ev > 0:
            equity_val = self.base_ev - self.net_debt_bn * 1e9
            self.shares_bn = equity_val / (self.market_price * 1e9) if equity_val > 0 else 1
        else:
            self.shares_bn = 1

        # Tax rate (standard)
        self.tax_rate = 0.21  # US federal + state blended

        # Build projection assumptions from historical data
        self._build_projections()

    def _build_projections(self):
        """Derive 5-year projection assumptions from historical data."""
        cagr = self.hist_cagr

        # Bear/base/bull revenue growth schedules (converge over 5 years)
        def growth_schedule(peak, terminal):
            step = (terminal - peak) / 4
            return [round(peak + step * i, 4) for i in range(5)]

        # Base: historical CAGR converging to 3% terminal
        base_peak = min(cagr, 0.30)   # cap at 30% for projection
        self.base_growth = growth_schedule(base_peak, 0.03)

        # Bull: 20% above base peak
        bull_peak = min(base_peak * 1.20, 0.40)
        self.bull_growth = growth_schedule(bull_peak, 0.035)

        # Bear: 30% below base peak
        bear_peak = max(base_peak * 0.70, 0.01)
        self.bear_growth = growth_schedule(bear_peak, 0.02)

        # Margin assumptions (converge to long-run from current)
        gm = max(self.gross_margin, 5) / 100
        ebitda_pct = self.ebitda_bn / (self.revenue_bn * 1e9) if self.revenue_bn > 0 else 0.15
        ebitda_pct = max(ebitda_pct, 0.05)
        da_pct  = max((self.ebitda_bn - self.ebit_margin/100 * self.revenue_bn) / (self.revenue_bn * 1e9), 0.03) if self.revenue_bn > 0 else 0.05
        da_pct  = min(da_pct, 0.12)
        sga_pct = max(gm - ebitda_pct, 0.05)
        capex_pct = max((-self.fcf_bn + self.ebitda_bn - (ebitda_pct - da_pct) * self.revenue_bn) / (self.revenue_bn * 1e9) if self.revenue_bn > 0 else 0.04, 0.02)
        capex_pct = min(capex_pct, 0.15)

        self.gm_pct    = round(gm, 4)
        self.ebitda_pct= round(ebitda_pct, 4)
        self.da_pct    = round(da_pct, 4)
        self.sga_pct   = round(sga_pct, 4)
        self.capex_pct = round(capex_pct, 4)
        self.cogs_pct  = round(1 - gm, 4)

        # NWC assumptions (typical ratios if not available)
        self.dso = 45.0   # days sales outstanding
        self.dih = 60.0   # days inventory held
        self.dpo = 40.0   # days payable outstanding
        self.accrued_pct = 0.05
        self.other_cl_pct = 0.02
        self.prepaids_pct = 0.02

        # Build 5-year projections (base scenario)
        rev = self.revenue_bn * 1e9  # in absolute $ (millions)
        self.proj_years   = list(range(2, 8))  # relative years
        self.proj_revenue = []
        self.proj_ebitda  = []
        self.proj_ebit    = []
        self.proj_ebiat   = []
        self.proj_da      = []
        self.proj_capex   = []
        self.proj_nwc_ch  = []
        self.proj_ufcf    = []

        prev_nwc = rev * 0.18  # initial NWC estimate

        for i, g in enumerate(self.base_growth):
            rev = rev * (1 + g)
            ebitda = rev * self.ebitda_pct
            da     = rev * self.da_pct
            ebit   = ebitda - da
            ebiat  = ebit * (1 - self.tax_rate)
            capex  = -rev * self.capex_pct
            nwc    = rev * 0.18
            dnwc   = -(nwc - prev_nwc)
            ufcf   = ebiat + da + capex + dnwc
            prev_nwc = nwc

            self.proj_revenue.append(round(rev / 1e6, 2))  # $M
            self.proj_ebitda.append(round(ebitda / 1e6, 2))
            self.proj_ebit.append(round(ebit / 1e6, 2))
            self.proj_ebiat.append(round(ebiat / 1e6, 2))
            self.proj_da.append(round(da / 1e6, 2))
            self.proj_capex.append(round(capex / 1e6, 2))
            self.proj_nwc_ch.append(round(dnwc / 1e6, 2))
            self.proj_ufcf.append(round(ufcf / 1e6, 2))

        # Last year values for terminal value
        self.terminal_ebitda = self.proj_ebitda[-1]
        self.terminal_ufcf   = self.proj_ufcf[-1]

    @staticmethod
    def _s(v) -> Optional[float]:
        if v is None: return None
        try:
            f = float(v)
            return None if math.isnan(f) or math.isinf(f) else f
        except: return None


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

class DCFExcelBuilder:

    def __init__(self, data: ReportData):
        self.d  = data
        self.wb = Workbook()

    def build(self, output_path: Path) -> Path:
        """Build the complete workbook and save."""
        ws_dcf  = self.wb.active; ws_dcf.title  = "DCF"
        ws_nwc  = self.wb.create_sheet("NWC")
        ws_wacc = self.wb.create_sheet("WACC")
        ws_as1  = self.wb.create_sheet("AS1")
        ws_as2  = self.wb.create_sheet("AS2")

        self._build_dcf(ws_dcf)
        self._build_nwc(ws_nwc)
        self._build_wacc(ws_wacc)
        self._build_as1(ws_as1)
        self._build_as2(ws_as2)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        return output_path

    # ── Column setup helper ─────────────────────────────────────────────────
    def _setup_cols(self, ws, widths: dict):
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.sheet_view.showGridLines = False

    # ── Common title block ──────────────────────────────────────────────────
    def _title_block(self, ws, subtitle: str, col_end: str = "N"):
        d = self.d
        ws.merge_cells(f"A1:{col_end}1")
        _set(ws, 1, 1, d.ticker, fill=_fill(DARK_BLUE),
             font=_font(bold=True, color=WHITE, size=14), align=_left())
        ws.merge_cells(f"A2:{col_end}2")
        _set(ws, 2, 1, subtitle, fill=_fill(DARK_BLUE),
             font=_font(bold=True, color=WHITE, size=11), align=_left())
        ws.merge_cells(f"A3:{col_end}3")
        meta = f"Lifecycle: {d.lifecycle_stage.replace('_',' ').title()}  |  " \
               f"Conviction: {d.conviction_score}  |  " \
               f"Tier: {d.position_tier.replace('_',' ').upper()}  |  " \
               f"Generated by Aletheia"
        _set(ws, 3, 1, meta, fill=_fill(DARK_BLUE),
             font=_font(color=GOLD, size=9, italic=True), align=_left())
        for r in [1, 2, 3]:
            ws.row_dimensions[r].height = 18

    # ── DCF SHEET ────────────────────────────────────────────────────────────
    def _build_dcf(self, ws):
        d = self.d
        self._setup_cols(ws, {
            "A":36, "B":3, "C":11,
            "D":11, "E":11, "F":11,   # hist years or empty
            "G":8,                     # CAGR
            "H":11,"I":11,"J":11,"K":11,"L":11,  # proj years
            "M":8,                     # proj CAGR
            "N":10,
        })
        self._title_block(ws, "Discounted Cash Flow Analysis")

        # --- Period header ---
        r = 4
        _hdr_row(ws, r, "Historical (Latest Year)", 4, 6)
        _set(ws, r, 7, "CAGR",
             fill=_fill(MID_BLUE), font=_hdr(), align=_center())
        _hdr_row(ws, r, "5-Year Projection (Base Case)", 8, 12)
        _set(ws, r, 13, "CAGR",
             fill=_fill(MID_BLUE), font=_hdr(), align=_center())

        r = 5
        base_year = 2024
        proj_yrs  = [base_year+1+i for i in range(5)]
        _set(ws, r, 1, "Assumptions: Base Scenario  |  Mid-Year Convention: N",
             fill=_fill(LIGHT_BLUE), font=_font(bold=True, size=8), align=_left())
        _set(ws, r, 4, base_year,
             fill=_fill(LIGHT_BLUE), font=_font(bold=True, size=8), align=_center(), fmt='0')
        for i, yr in enumerate(proj_yrs):
            _set(ws, r, 8+i, yr,
                 fill=_fill(LIGHT_BLUE), font=_font(bold=True, size=8),
                 align=_center(), fmt='0')

        # Historical anchors (latest year)
        rev_hist   = d.revenue_bn * 1e3     # $M
        ebitda_hist= d.ebitda_bn * 1e3
        ebit_hist  = ebitda_hist * (1 - d.da_pct / max(d.ebitda_pct,0.01))
        da_hist    = ebitda_hist - ebit_hist
        cogs_hist  = rev_hist * d.cogs_pct
        sga_hist   = rev_hist * d.sga_pct
        tax_hist   = ebit_hist * d.tax_rate
        ebiat_hist = ebit_hist - tax_hist
        capex_hist = -rev_hist * d.capex_pct

        hist_cagr = d.hist_cagr
        proj_cagr = (d.proj_revenue[-1] / (rev_hist/1e3))**(1/5) - 1 if rev_hist > 0 else 0
        proj_ebiat_cagr = (d.proj_ebiat[-1] / (ebiat_hist/1e3))**(1/5) - 1 if ebiat_hist != 0 else 0

        def prow(ws, row, label, h_val, p_vals, cagr_h=None, cagr_p=None,
                 fmt=FMT_DOLLAR, bold=False, bg=WHITE, is_input=False):
            """Write a P&L row: label | hist | blank | proj x5 | cagr."""
            _set(ws, row, 1, label, fill=_fill(bg),
                 font=_font(bold=bold, size=8), align=_left())
            _set(ws, row, 2, None, fill=_fill(bg))
            _set(ws, row, 3, None, fill=_fill(bg))
            fn = _input() if is_input else _font(size=8)
            if h_val is not None:
                _set(ws, row, 4, h_val, fill=_fill(bg), font=fn, align=_right(), fmt=fmt)
            _set(ws, row, 5, None, fill=_fill(bg))
            _set(ws, row, 6, None, fill=_fill(bg))
            if cagr_h is not None:
                _set(ws, row, 7, cagr_h, fill=_fill(bg),
                     font=_font(size=8), align=_center(), fmt=FMT_PCT1)
            for i, v in enumerate(p_vals):
                _set(ws, row, 8+i, v, fill=_fill(bg),
                     font=_link() if not is_input else _input(), align=_right(), fmt=fmt)
            if cagr_p is not None:
                _set(ws, row, 13, cagr_p, fill=_fill(bg),
                     font=_font(size=8), align=_center(), fmt=FMT_PCT1)

        r = 6
        # Income Statement
        _hdr_row(ws, r, "Income Statement ($M)", 1, 13); r += 1

        prow(ws, r, "Revenue", rev_hist, d.proj_revenue,
             cagr_h=hist_cagr, cagr_p=proj_cagr,
             bg=VERY_LIGHT, is_input=True); r += 1
        prow(ws, r, "   % growth", None,
             [g for g in d.base_growth],
             fmt=FMT_PCT1, bg=WHITE); r += 1
        prow(ws, r, "Cost of Goods Sold", cogs_hist,
             [v*d.cogs_pct for v in d.proj_revenue],
             bg=VERY_LIGHT, is_input=True); r += 1
        prow(ws, r, "Gross Profit",
             rev_hist - cogs_hist,
             [v*(1-d.cogs_pct) for v in d.proj_revenue],
             bg=WHITE, bold=False); r += 1
        prow(ws, r, "   % margin", None,
             [1-d.cogs_pct]*5, fmt=FMT_PCT1, bg=VERY_LIGHT); r += 1
        prow(ws, r, "SG&A", sga_hist,
             [v*d.sga_pct for v in d.proj_revenue],
             bg=WHITE, is_input=True); r += 1
        prow(ws, r, "EBITDA", ebitda_hist, d.proj_ebitda,
             bg=LIGHT_BLUE, bold=True); r += 1
        prow(ws, r, "   % margin", None,
             [e/r_ for e,r_ in zip(d.proj_ebitda, d.proj_revenue)],
             fmt=FMT_PCT1, bg=VERY_LIGHT); r += 1
        prow(ws, r, "Depreciation & Amortization", da_hist, d.proj_da,
             bg=WHITE); r += 1
        prow(ws, r, "EBIT", ebit_hist, d.proj_ebit,
             bg=LIGHT_BLUE, bold=True); r += 1
        prow(ws, r, "   % margin", None,
             [e/r_ for e,r_ in zip(d.proj_ebit, d.proj_revenue)],
             fmt=FMT_PCT1, bg=VERY_LIGHT); r += 1
        prow(ws, r, f"Taxes (@{d.tax_rate:.0%})", tax_hist,
             [e*d.tax_rate for e in d.proj_ebit],
             bg=WHITE, is_input=True); r += 1
        prow(ws, r, "EBIAT (NOPAT)", ebiat_hist, d.proj_ebiat,
             cagr_h=None, cagr_p=proj_ebiat_cagr,
             bg=LIGHT_BLUE, bold=True); r += 2

        # FCF Build
        _hdr_row(ws, r, "Free Cash Flow Build ($M)", 1, 13); r += 1
        prow(ws, r, "Plus: Depreciation & Amortization",
             da_hist, d.proj_da, bg=VERY_LIGHT); r += 1
        prow(ws, r, "Less: Capital Expenditures",
             capex_hist, d.proj_capex, bg=WHITE); r += 1
        prow(ws, r, "Less: Change in Net Working Capital",
             None, d.proj_nwc_ch, bg=VERY_LIGHT); r += 1
        prow(ws, r, "Unlevered Free Cash Flow",
             None, d.proj_ufcf,
             bg=LIGHT_BLUE, bold=True); r += 1

        # Discount factors
        wacc = d.wacc
        disc = [(1/(1+wacc))**t for t in range(1, 6)]
        pv_fcf = [f*u for f,u in zip(disc, d.proj_ufcf)]

        prow(ws, r, f"   WACC = {wacc:.2%}", None,
             [None]*5, bg=WHITE)
        _set(ws, r, 4, wacc, fill=_fill(WHITE), font=_input(), align=_right(), fmt=FMT_PCT1)
        r += 1
        prow(ws, r, "   Discount Period", None,
             [1,2,3,4,5], fmt="0.0", bg=VERY_LIGHT); r += 1
        prow(ws, r, "   Discount Factor", None,
             disc, fmt=FMT_DEC4, bg=WHITE); r += 1
        prow(ws, r, "   PV of Free Cash Flow", None,
             pv_fcf, bg=LIGHT_BLUE, bold=True); r += 2

        # ── Valuation Summary: Exit Multiple Method ──────────────────────────
        cum_pv   = sum(pv_fcf)
        tv_ebitda= d.terminal_ebitda   # $M
        tv_mult  = round(max(d.justified_ev_ebitda, 6.0), 1)  # use justified as base
        tv_mult  = min(tv_mult, 20.0)
        tv_exit  = tv_ebitda * tv_mult
        tv_disc  = tv_exit * disc[-1]
        ev_mult  = cum_pv + tv_disc
        debt_m   = d.net_debt_bn * 1e3  # $M
        equity_m = ev_mult - debt_m
        shares   = d.shares_bn * 1e3    # thousands to millions

        _hdr_row(ws, r, "Valuation — Exit Multiple Method", 1, 7)
        _hdr_row(ws, r, "Implied Equity Value & Share Price", 8, 13); r += 1

        def vline(ws, row, label, val, col=4, fmt=FMT_DOLLAR, bold=False, bg=WHITE):
            _set(ws, row, 1, label, fill=_fill(bg), font=_font(bold=bold, size=8), align=_left())
            _set(ws, row, col, val, fill=_fill(bg),
                 font=_font(bold=bold, size=8), align=_right(), fmt=fmt)

        vline(ws, r, "Cumulative PV of FCFs", cum_pv, bg=VERY_LIGHT); r_bridge = r
        r += 1
        vline(ws, r, f"Terminal Year EBITDA ({proj_yrs[-1]}E)", tv_ebitda, bg=WHITE); r += 1
        vline(ws, r, "Exit Multiple", tv_mult, fmt=FMT_MULT, bg=VERY_LIGHT); r += 1
        vline(ws, r, "   Terminal Value", tv_exit, bg=WHITE); r += 1
        vline(ws, r, "Discount Factor", disc[-1], fmt=FMT_DEC4, bg=VERY_LIGHT); r += 1
        vline(ws, r, "   PV of Terminal Value", tv_disc, bg=WHITE); r += 1
        vline(ws, r, "   % of Enterprise Value", tv_disc/ev_mult if ev_mult else 0,
              fmt=FMT_PCT1, bg=VERY_LIGHT); r += 2
        vline(ws, r, "   Enterprise Value", ev_mult, bold=True, bg=LIGHT_BLUE); r += 1

        # Bridge
        bridge_row = r_bridge
        vline(ws, bridge_row,   "Enterprise Value", ev_mult,  col=10, bg=VERY_LIGHT); bridge_row += 1
        vline(ws, bridge_row,   "Less: Net Debt", -debt_m,    col=10, bg=WHITE); bridge_row += 1
        vline(ws, bridge_row,   "Less: Minority Interest", 0, col=10, bg=VERY_LIGHT); bridge_row += 1
        vline(ws, bridge_row,   "Plus: Cash",
              max(-debt_m, 0), col=10, bg=WHITE); bridge_row += 2
        vline(ws, bridge_row,   "Implied Equity Value", equity_m, col=10,
              bold=True, bg=LIGHT_BLUE); bridge_row += 2
        vline(ws, bridge_row,   "Shares Outstanding (M)", shares, col=10,
              fmt=FMT_INT, bg=VERY_LIGHT); bridge_row += 2
        price = equity_m / shares if shares else 0
        vline(ws, bridge_row,   "   Intrinsic Share Price", price, col=10,
              bold=True, bg=LIGHT_BLUE, fmt=FMT_DOLLAR2)

        # Aletheia scenario comparison
        r += 2
        _hdr_row(ws, r, "Aletheia Three-Scenario DCF Comparison", 1, 13); r += 1
        scenarios = [("Bear", d.bear_iv, d.bear_mos, "2E5FA3"),
                     ("Base", d.base_iv, d.base_mos, "1F3864"),
                     ("Bull", d.bull_iv, d.bull_mos, "1a6b3c")]
        for sc, iv, mos, col_hex in scenarios:
            _set(ws, r, 1, sc, fill=_fill(col_hex),
                 font=_font(bold=True, color=WHITE, size=9), align=_center())
            _set(ws, r, 2, f"IV: ${iv:.2f}", fill=_fill(col_hex),
                 font=_font(color=WHITE, size=9), align=_center())
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            _set(ws, r, 6, f"MoS: {mos:+.1%}", fill=_fill(col_hex),
                 font=_font(color=WHITE, size=9), align=_center())
            ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
            if d.market_price:
                _set(ws, r, 9, f"Market: ${d.market_price:.2f}", fill=_fill(col_hex),
                     font=_font(color=WHITE, size=9), align=_center())
                ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
            r += 1

        # ── Sensitivity: Exit Multiple ────────────────────────────────────────
        r += 2
        _hdr_row(ws, r, "Sensitivity — Enterprise Value ($M)  |  Exit Multiple Method", 1, 9); r += 1

        exit_mults  = [tv_mult-2, tv_mult-1, tv_mult, tv_mult+1, tv_mult+2]
        wacc_range  = [wacc-0.02, wacc-0.01, wacc, wacc+0.01, wacc+0.02]

        _set(ws, r, 1, "WACC \\ Exit Multiple",
             fill=_fill(LIGHT_BLUE), font=_font(bold=True, size=8), align=_center())
        for i, em in enumerate(exit_mults):
            _set(ws, r, 2+i, f"{em:.1f}x",
                 fill=_fill(LIGHT_BLUE), font=_font(bold=True, size=8), align=_center())
        r += 1

        for wi, wc in enumerate(wacc_range):
            _set(ws, r, 1, wc, fill=_fill(LIGHT_BLUE),
                 font=_font(bold=True, size=8), align=_center(), fmt=FMT_PCT1)
            df_t = (1/(1+wc))**5
            cum_wc = sum((d.proj_ufcf[i])*(1/(1+wc))**(i+1) for i in range(5))
            for ei, em in enumerate(exit_mults):
                tv_wc = tv_ebitda * em * df_t
                ev_wc = cum_wc + tv_wc
                is_base = (wi == 2 and ei == 2)
                bg = SENS_GOLD if is_base else (VERY_LIGHT if wi % 2 == 0 else WHITE)
                _set(ws, r, 2+ei, ev_wc,
                     fill=_fill(bg),
                     font=_font(bold=is_base, color=RED_DARK if is_base else FORMULA_BLK, size=8),
                     align=_center(), fmt='$#,##0')
            r += 1

        # ── Sensitivity: Growth Rate ─────────────────────────────────────────
        r += 2
        _hdr_row(ws, r, "Sensitivity — Enterprise Value ($M)  |  Perpetuity Growth Rate Method", 1, 9); r += 1

        tv_fcf     = d.terminal_ufcf
        growth_g   = 0.03
        tv_gordon  = tv_fcf * (1+growth_g) / (wacc - growth_g) if wacc > growth_g else 0
        tv_gordon_d= tv_gordon * disc[-1]
        ev_gordon  = cum_pv + tv_gordon_d

        growth_rates = [0.01, 0.02, 0.03, 0.035, 0.04]
        _set(ws, r, 1, "WACC \\ Growth Rate",
             fill=_fill(LIGHT_BLUE), font=_font(bold=True, size=8), align=_center())
        for i, gr in enumerate(growth_rates):
            _set(ws, r, 2+i, gr, fill=_fill(LIGHT_BLUE),
                 font=_font(bold=True, size=8), align=_center(), fmt=FMT_PCT1)
        r += 1

        for wi, wc in enumerate(wacc_range):
            _set(ws, r, 1, wc, fill=_fill(LIGHT_BLUE),
                 font=_font(bold=True, size=8), align=_center(), fmt=FMT_PCT1)
            df_t  = (1/(1+wc))**5
            cum_wc= sum((d.proj_ufcf[i])*(1/(1+wc))**(i+1) for i in range(5))
            for gi, gr in enumerate(growth_rates):
                if wc <= gr:
                    ev_g = None
                else:
                    tv_g = tv_fcf*(1+gr)/(wc-gr)
                    ev_g = cum_wc + tv_g*df_t
                is_base = (wi == 2 and gi == 2)
                bg = SENS_GOLD if is_base else (VERY_LIGHT if wi % 2 == 0 else WHITE)
                _set(ws, r, 2+gi, ev_g,
                     fill=_fill(bg),
                     font=_font(bold=is_base, color=RED_DARK if is_base else FORMULA_BLK, size=8),
                     align=_center(), fmt='$#,##0' if ev_g else None)
            r += 1

        # Reverse DCF callout
        r += 1
        _hdr_row(ws, r, "Reverse DCF — Market-Implied Growth", 1, 9); r += 1
        rdcf_items = [
            ("Historical Revenue CAGR", d.hist_cagr, FMT_PCT1),
            ("Market-Implied CAGR (10Y)", d.impl_cagr, FMT_PCT1),
            ("Implied / Historical Ratio", d.impl_cagr/d.hist_cagr if d.hist_cagr else 0, "0.0x"),
            ("Signal", d.rdcf_signal, None),
            ("Market EV/EBITDA", d.market_ev_ebitda, FMT_MULT),
            ("Justified EV/EBITDA", d.justified_ev_ebitda, FMT_MULT),
            ("ROIC-WACC Spread", d.roic_wacc_spread, "+0.0%;(0.0%);-"),
            ("Value Creation", d.value_creation, None),
        ]
        for i, (label, val, fmt) in enumerate(rdcf_items):
            bg = VERY_LIGHT if i % 2 == 0 else WHITE
            _set(ws, r, 1, label, fill=_fill(bg), font=_font(size=8), align=_left())
            if fmt:
                _set(ws, r, 4, val, fill=_fill(bg), font=_font(bold=True, size=8),
                     align=_right(), fmt=fmt)
            else:
                _set(ws, r, 4, str(val).replace("_", " ").title() if val else "",
                     fill=_fill(bg), font=_font(bold=True, size=8), align=_right())
            r += 1

    # ── NWC SHEET ────────────────────────────────────────────────────────────
    def _build_nwc(self, ws):
        d = self.d
        self._setup_cols(ws, {"A":34,"B":3,"C":11,"D":11,"E":11,"F":11,"G":11,"H":11,"I":11})
        self._title_block(ws, "Net Working Capital Schedule ($M)", col_end="I")

        r = 4
        base_year = 2024
        proj_yrs  = [base_year+1+i for i in range(5)]
        _hdr_row(ws, r, "Historical", 3, 5)
        _hdr_row(ws, r, "Projection Period", 6, 9)
        r += 1
        _set(ws, r, 3, base_year-1, fill=_fill(LIGHT_BLUE), font=_font(bold=True,size=8), align=_center(), fmt='0')
        _set(ws, r, 4, base_year-0, fill=_fill(LIGHT_BLUE), font=_font(bold=True,size=8), align=_center(), fmt='0')
        _set(ws, r, 5, base_year, fill=_fill(LIGHT_BLUE), font=_font(bold=True,size=8), align=_center(), fmt='0')
        for i, yr in enumerate(proj_yrs[:4]):
            _set(ws, r, 6+i, yr, fill=_fill(LIGHT_BLUE), font=_font(bold=True,size=8), align=_center(), fmt='0')
        r += 1

        rev_m   = d.revenue_bn * 1e3
        cogs_m  = rev_m * d.cogs_pct
        dso, dih, dpo = d.dso, d.dih, d.dpo

        proj_rev  = d.proj_revenue[:4]
        proj_cogs = [v*d.cogs_pct for v in proj_rev]

        def nwc_row(ws, row, label, hist_vals, proj_vals, fmt=FMT_DOLLAR, bold=False, bg=WHITE):
            _set(ws, row, 1, label, fill=_fill(bg), font=_font(bold=bold,size=8), align=_left())
            for i, v in enumerate(hist_vals):
                _set(ws, row, 3+i, v, fill=_fill(bg), font=_input(), align=_right(), fmt=fmt)
            for i, v in enumerate(proj_vals):
                _set(ws, row, 3+len(hist_vals)+i, v, fill=_fill(bg),
                     font=_link(), align=_right(), fmt=fmt)

        # Key drivers
        nwc_row(ws, r, "Revenue", [rev_m*0.85, rev_m*0.92, rev_m], proj_rev, bg=VERY_LIGHT); r += 1
        nwc_row(ws, r, "Cost of Goods Sold", [cogs_m*0.85, cogs_m*0.92, cogs_m], proj_cogs); r += 2

        _hdr_row(ws, r, "Current Assets", 1, 9); r += 1
        ar   = [v*dso/365 for v in [rev_m*0.85, rev_m*0.92, rev_m]]
        ar_p = [v*dso/365 for v in proj_rev]
        inv  = [v*dih/365 for v in [cogs_m*0.85, cogs_m*0.92, cogs_m]]
        inv_p= [v*dih/365 for v in proj_cogs]
        pre  = [v*d.prepaids_pct for v in [rev_m*0.85, rev_m*0.92, rev_m]]
        pre_p= [v*d.prepaids_pct for v in proj_rev]
        tca  = [a+b+c for a,b,c in zip(ar, inv, pre)]
        tca_p= [a+b+c for a,b,c in zip(ar_p, inv_p, pre_p)]

        nwc_row(ws, r, "Accounts Receivable", ar, ar_p, bg=VERY_LIGHT); r += 1
        nwc_row(ws, r, "Inventories", inv, inv_p); r += 1
        nwc_row(ws, r, "Prepaid & Other", pre, pre_p, bg=VERY_LIGHT); r += 1
        nwc_row(ws, r, "   Total Current Assets", tca, tca_p, bold=True, bg=LIGHT_BLUE); r += 2

        _hdr_row(ws, r, "Current Liabilities", 1, 9); r += 1
        ap   = [v*dpo/365 for v in [cogs_m*0.85, cogs_m*0.92, cogs_m]]
        ap_p = [v*dpo/365 for v in proj_cogs]
        acc  = [v*d.accrued_pct for v in [rev_m*0.85, rev_m*0.92, rev_m]]
        acc_p= [v*d.accrued_pct for v in proj_rev]
        ocl  = [v*d.other_cl_pct for v in [rev_m*0.85, rev_m*0.92, rev_m]]
        ocl_p= [v*d.other_cl_pct for v in proj_rev]
        tcl  = [a+b+c for a,b,c in zip(ap, acc, ocl)]
        tcl_p= [a+b+c for a,b,c in zip(ap_p, acc_p, ocl_p)]

        nwc_row(ws, r, "Accounts Payable", ap, ap_p, bg=VERY_LIGHT); r += 1
        nwc_row(ws, r, "Accrued Liabilities", acc, acc_p); r += 1
        nwc_row(ws, r, "Other Current Liabilities", ocl, ocl_p, bg=VERY_LIGHT); r += 1
        nwc_row(ws, r, "   Total Current Liabilities", tcl, tcl_p, bold=True, bg=LIGHT_BLUE); r += 2

        nwc  = [a-b for a,b in zip(tca, tcl)]
        nwc_p= [a-b for a,b in zip(tca_p, tcl_p)]
        nwc_row(ws, r, "   Net Working Capital", nwc, nwc_p, bold=True, bg=LIGHT_BLUE); r += 1
        nwc_pct = [n/v for n,v in zip(nwc, [rev_m*0.85, rev_m*0.92, rev_m])]
        nwc_pct_p=[n/v for n,v in zip(nwc_p, proj_rev)]
        nwc_row(ws, r, "   % of Revenue", nwc_pct, nwc_pct_p, fmt=FMT_PCT1); r += 2

        all_nwc = nwc + nwc_p
        dnwc = [None, None] + [-(all_nwc[i]-all_nwc[i-1]) for i in range(1, len(all_nwc))]
        nwc_row(ws, r, "   Change in NWC (+ = source)", dnwc[:3], dnwc[3:7], bold=True, bg=LIGHT_BLUE); r += 2

        _hdr_row(ws, r, "Assumptions", 1, 9); r += 1
        for label, val, fmt in [
            ("Days Sales Outstanding (DSO)", dso, "0.0"),
            ("Days Inventory Held (DIH)",    dih, "0.0"),
            ("Days Payable Outstanding (DPO)",dpo, "0.0"),
            ("Prepaids (% of Revenue)",    d.prepaids_pct, FMT_PCT1),
            ("Accrued Liabilities (% of Revenue)", d.accrued_pct, FMT_PCT1),
            ("Other CL (% of Revenue)",    d.other_cl_pct, FMT_PCT1),
        ]:
            bg = VERY_LIGHT if r % 2 == 0 else WHITE
            _set(ws, r, 1, label, fill=_fill(bg), font=_font(size=8), align=_left())
            _set(ws, r, 4, val, fill=_fill(bg), font=_input(), align=_center(), fmt=fmt)
            r += 1

    # ── WACC SHEET ───────────────────────────────────────────────────────────
    def _build_wacc(self, ws):
        d = self.d
        self._setup_cols(ws, {"A":34,"B":3,"C":14,"D":12,"E":12,"F":12,"G":12,"H":12,"I":12})
        self._title_block(ws, "Weighted Average Cost of Capital (WACC) Analysis", col_end="I")

        r = 4
        _hdr_row(ws, r, "WACC Build", 1, 4); r += 1

        def wacc_line(ws, row, label, val, fmt=FMT_PCT1, bold=False, bg=WHITE):
            _set(ws, row, 1, label, fill=_fill(bg), font=_font(bold=bold,size=8), align=_left())
            _set(ws, row, 3, val, fill=_fill(bg),
                 font=_input() if not bold else _font(bold=True,size=8),
                 align=_center(), fmt=fmt)

        ke = d.rf + d.beta * d.erp
        kd_pretax = 0.055   # assumed
        kd = kd_pretax * (1 - d.tax_rate)
        lev  = 0.30
        wacc_calc = lev * kd + (1-lev) * ke

        wacc_line(ws, r, "Target Capital Structure", None, bg=LIGHT_BLUE)
        _set(ws, r, 1, "Target Capital Structure", fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=8), align=_left()); r += 1
        wacc_line(ws, r, "Debt / Total Capital", lev, bg=VERY_LIGHT); r += 1
        wacc_line(ws, r, "Equity / Total Capital", 1-lev, bg=WHITE); r += 2

        _set(ws, r, 1, "Cost of Debt", fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=8), align=_left()); r += 1
        wacc_line(ws, r, "Pre-tax Cost of Debt", kd_pretax, bg=VERY_LIGHT); r += 1
        wacc_line(ws, r, "Tax Rate", d.tax_rate, bg=WHITE); r += 1
        wacc_line(ws, r, "   After-tax Cost of Debt", kd, bg=LIGHT_BLUE, bold=True); r += 2

        _set(ws, r, 1, "Cost of Equity (CAPM)", fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=8), align=_left()); r += 1
        wacc_line(ws, r, "Risk-free Rate (10Y Treasury)", d.rf, bg=VERY_LIGHT); r += 1
        wacc_line(ws, r, "Equity Risk Premium (Damodaran)", d.erp, bg=WHITE); r += 1
        wacc_line(ws, r, f"Beta (5Y Weekly vs SPY)", d.beta, fmt="0.000", bg=VERY_LIGHT); r += 1
        wacc_line(ws, r, "   Cost of Equity (CAPM)", ke, bg=LIGHT_BLUE, bold=True); r += 2

        _set(ws, r, 1, "   WACC", fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=9), align=_left())
        _set(ws, r, 3, d.wacc, fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=12,color=GOLD), align=_center(), fmt=FMT_PCT1)
        _set(ws, r, 4, f"(Aletheia live: {d.wacc:.2%}  |  CAPM rebuild: {wacc_calc:.2%})",
             fill=_fill(LIGHT_BLUE), font=_font(size=8,italic=True,color="666666"), align=_left())
        r += 3

        # WACC Sensitivity
        _hdr_row(ws, r, "WACC Sensitivity  (Debt/Total Capital × Pre-tax Cost of Debt)", 1, 8); r += 1
        kd_range  = [0.040, 0.050, 0.060, 0.070, 0.080]
        lev_range = [0.10,  0.20,  0.30,  0.40,  0.50]

        _set(ws, r, 1, "D/(D+E) \\ Kd", fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=8), align=_center())
        for i, kd_v in enumerate(kd_range):
            _set(ws, r, 2+i, kd_v, fill=_fill(LIGHT_BLUE),
                 font=_font(bold=True,size=8), align=_center(), fmt=FMT_PCT1)
        r += 1

        for li, lv in enumerate(lev_range):
            _set(ws, r, 1, lv, fill=_fill(LIGHT_BLUE),
                 font=_font(bold=True,size=8), align=_center(), fmt=FMT_PCT1)
            for ki, kd_v in enumerate(kd_range):
                wc = lv * kd_v*(1-d.tax_rate) + (1-lv)*ke
                is_base = (li == 2 and ki == 2)
                bg = SENS_GOLD if is_base else (VERY_LIGHT if li%2==0 else WHITE)
                _set(ws, r, 2+ki, wc, fill=_fill(bg),
                     font=_font(bold=is_base, color=RED_DARK if is_base else FORMULA_BLK, size=8),
                     align=_center(), fmt=FMT_PCT1)
            r += 1

        r += 1
        for note in [
            f"(1) Risk-free rate: {d.rf:.2%} — live 10Y US Treasury from Aletheia pipeline",
            f"(2) Beta: {d.beta:.3f} — 5-year weekly regression vs SPY",
            "(3) ERP: 6.62% — Damodaran Implied ERP",
        ]:
            _set(ws, r, 1, note, font=_font(size=8,italic=True,color="666666"), align=_left())
            r += 1

    # ── AS1 SHEET ────────────────────────────────────────────────────────────
    def _build_as1(self, ws):
        d = self.d
        self._setup_cols(ws, {"A":40,"B":3,"C":10,"D":8,**{get_column_letter(c):9 for c in range(5,15)}})
        self._title_block(ws, "Assumptions — Income Statement & Cash Flow", col_end="N")

        base_year = 2024
        proj_yrs  = [base_year+1+i for i in range(5)]
        r = 4
        _hdr_row(ws, r, "Projection Period", 5, 14); r += 1
        for i, yr in enumerate(proj_yrs):
            _set(ws, r, 5+i, yr, fill=_fill(LIGHT_BLUE),
                 font=_font(bold=True,size=8), align=_center(), fmt='0')
        _set(ws, r, 4, "Scenario#", fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=8), align=_center()); r += 1

        def as_section(ws, row, title, items):
            _hdr_row(ws, row, title, 1, 14); row += 1
            for i, (label, num, rates, fmt) in enumerate(items):
                is_hdr = num is None
                bg = LIGHT_BLUE if is_hdr else (VERY_LIGHT if row%2==0 else WHITE)
                _set(ws, row, 1, label, fill=_fill(bg),
                     font=_font(bold=is_hdr,size=8), align=_left())
                if num is not None:
                    _set(ws, row, 4, num, fill=_fill(bg), font=_input(), align=_center())
                for j, rate in enumerate(rates):
                    _set(ws, row, 5+j, rate, fill=_fill(bg),
                         font=_input(), align=_center(), fmt=fmt)
                row += 1
            return row + 1

        g = d.base_growth
        cagr = d.hist_cagr

        r = as_section(ws, r, "Revenue Growth (% YoY)", [
            ("Revenue Growth — Live",  None, g, FMT_PCT1),
            ("   Base",    1, g, FMT_PCT1),
            ("   Upside",  2, [min(x*1.3, 0.5) for x in g], FMT_PCT1),
            ("   Downside",3, [max(x*0.6, 0.01) for x in g], FMT_PCT1),
        ])
        r = as_section(ws, r, "Gross Margin (% of Revenue)", [
            ("Gross Margin — Live", None, [d.gm_pct]*5, FMT_PCT1),
            ("   Base",    1, [d.gm_pct]*5, FMT_PCT1),
            ("   Upside",  2, [min(d.gm_pct+0.02, 0.99)]*5, FMT_PCT1),
            ("   Downside",3, [max(d.gm_pct-0.03, 0.10)]*5, FMT_PCT1),
        ])
        r = as_section(ws, r, "EBITDA Margin (% of Revenue)", [
            ("EBITDA Margin — Live", None, [d.ebitda_pct]*5, FMT_PCT1),
            ("   Base",    1, [d.ebitda_pct]*5, FMT_PCT1),
            ("   Upside",  2, [min(d.ebitda_pct+0.03, 0.60)]*5, FMT_PCT1),
            ("   Downside",3, [max(d.ebitda_pct-0.05, 0.02)]*5, FMT_PCT1),
        ])
        r = as_section(ws, r, "D&A (% of Revenue)", [
            ("D&A — Live", None, [d.da_pct]*5, FMT_PCT1),
            ("   Base",    1, [d.da_pct]*5, FMT_PCT1),
        ])
        r = as_section(ws, r, "CapEx (% of Revenue)", [
            ("CapEx — Live", None, [d.capex_pct]*5, FMT_PCT1),
            ("   Base",      1, [d.capex_pct]*5, FMT_PCT1),
            ("   Upside",    2, [max(d.capex_pct-0.01, 0.01)]*5, FMT_PCT1),
            ("   Downside",  3, [d.capex_pct+0.02]*5, FMT_PCT1),
        ])
        r = as_section(ws, r, "Tax Rate", [
            ("Tax Rate", 1, [d.tax_rate]*5, FMT_PCT1),
        ])

    # ── AS2 SHEET ────────────────────────────────────────────────────────────
    def _build_as2(self, ws):
        d = self.d
        self._setup_cols(ws, {"A":40,"B":3,"C":10,"D":8,**{get_column_letter(c):9 for c in range(5,15)}})
        self._title_block(ws, "Assumptions — Balance Sheet (NWC Drivers)", col_end="N")

        base_year = 2024
        proj_yrs  = [base_year+1+i for i in range(5)]
        r = 4
        _hdr_row(ws, r, "Projection Period", 5, 14); r += 1
        for i, yr in enumerate(proj_yrs):
            _set(ws, r, 5+i, yr, fill=_fill(LIGHT_BLUE),
                 font=_font(bold=True,size=8), align=_center(), fmt='0')
        _set(ws, r, 4, "Scenario#", fill=_fill(LIGHT_BLUE),
             font=_font(bold=True,size=8), align=_center()); r += 1

        def as2_section(ws, row, title, items):
            _hdr_row(ws, row, title, 1, 14); row += 1
            for label, num, vals, fmt in items:
                is_hdr = num is None
                bg = LIGHT_BLUE if is_hdr else (VERY_LIGHT if row%2==0 else WHITE)
                _set(ws, row, 1, label, fill=_fill(bg),
                     font=_font(bold=is_hdr,size=8), align=_left())
                if num is not None:
                    _set(ws, row, 4, num, fill=_fill(bg), font=_input(), align=_center())
                for j, v in enumerate(vals):
                    _set(ws, row, 5+j, v, fill=_fill(bg),
                         font=_input(), align=_center(), fmt=fmt)
                row += 1
            return row + 1

        r = as2_section(ws, r, "Days Sales Outstanding (DSO)", [
            ("DSO — Live",  None, [d.dso]*5, "0.0"),
            ("   Base",     1,    [d.dso]*5, "0.0"),
            ("   Upside",   2,    [d.dso*0.9]*5, "0.0"),
            ("   Downside", 3,    [d.dso*1.1]*5, "0.0"),
        ])
        r = as2_section(ws, r, "Days Inventory Held (DIH)", [
            ("DIH — Live",  None, [d.dih]*5, "0.0"),
            ("   Base",     1,    [d.dih]*5, "0.0"),
            ("   Upside",   2,    [d.dih*0.85]*5, "0.0"),
            ("   Downside", 3,    [d.dih*1.15]*5, "0.0"),
        ])
        r = as2_section(ws, r, "Days Payable Outstanding (DPO)", [
            ("DPO — Live",  None, [d.dpo]*5, "0.0"),
            ("   Base",     1,    [d.dpo]*5, "0.0"),
            ("   Upside",   2,    [d.dpo*1.1]*5, "0.0"),
            ("   Downside", 3,    [d.dpo*0.9]*5, "0.0"),
        ])
        r = as2_section(ws, r, "Prepaids & Other CA (% of Revenue)", [
            ("Prepaids — Live", None, [d.prepaids_pct]*5, FMT_PCT1),
            ("   Base",         1,    [d.prepaids_pct]*5, FMT_PCT1),
        ])
        r = as2_section(ws, r, "Accrued Liabilities (% of Revenue)", [
            ("Accrued — Live", None, [d.accrued_pct]*5, FMT_PCT1),
            ("   Base",        1,    [d.accrued_pct]*5, FMT_PCT1),
        ])


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def export_ticker(ticker: str) -> Optional[Path]:
    ticker = ticker.upper()
    report_path = REPORT_DIR / f"{ticker}_report.json"
    if not report_path.exists():
        print(f"  {ticker}: no report found at {report_path}")
        return None
    try:
        report = json.loads(report_path.read_text())
        data   = ReportData(report, ticker)
        out    = OUTPUT_DIR / f"{ticker}_DCF_Model.xlsx"
        builder= DCFExcelBuilder(data)
        path   = builder.build(out)
        print(f"  ✓ {ticker}: {path}")
        return path
    except Exception as e:
        import traceback
        print(f"  ✗ {ticker}: {e}")
        traceback.print_exc()
        return None


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate DCF Excel for Aletheia universe tickers")
    parser.add_argument("--ticker", nargs="+", help="Ticker(s) to export")
    parser.add_argument("--all",    action="store_true", help="Export all available tickers")
    args = parser.parse_args()

    if args.all:
        tickers = [p.stem.replace("_report", "")
                   for p in REPORT_DIR.glob("*_report.json")]
    elif args.ticker:
        tickers = args.ticker
    else:
        parser.print_help(); sys.exit(1)

    print(f"\nExporting {len(tickers)} ticker(s)...\n")
    results = [export_ticker(t) for t in tickers]
    success = [r for r in results if r]
    print(f"\n{'='*50}")
    print(f"Done: {len(success)}/{len(tickers)} exported to {OUTPUT_DIR}/")
